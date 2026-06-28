"""SIMS depth-profile parser (.csv/.tsv/.xlsx). Port of parser.importSIMS.

Handles two layouts:
  * shared-depth — column 0 is the depth axis, columns 1.. are concentrations;
  * paired — each element has its own (depth, concentration) column pair.

For paired files whose elements share an identical depth grid, no interpolation
is done. Otherwise a union depth grid (finest positive step over the full
range) is built and each element is linearly interpolated onto it (NaN outside
its measured range), matching MATLAB's ``interp1(..., 'linear', NaN)``.
"""

from __future__ import annotations

import math
import re
from collections.abc import Sequence
from pathlib import Path
from typing import Any

import numpy as np

from quantized.datastruct import DataStruct

__all__ = ["import_sims", "is_sims_file"]

_COMMENT_CHARS = "#%"
_DELIM_CANDIDATES = (",", "\t", ";", " ")
_EXCEL_EXTS = {".xlsx", ".xls", ".xlsm", ".xlsb", ".ods"}

# SIMS exports share .csv/.tsv/.xlsx with generic tables, so detect them by a
# vendor banner ("SIMS", "Evans Analytical", "Drawn Curves") or, failing that,
# the structural fingerprint of a depth axis carrying concentration units.
_SIMS_WORD_RE = re.compile(r"\bsims\b", re.IGNORECASE)
_SIMS_PHRASE_MARKERS = ("evans analytical", "drawn curves", "secondary ion")


def _sims_signals(text: str) -> bool:
    if _SIMS_WORD_RE.search(text):
        return True
    low = text.lower()
    if any(m in low for m in _SIMS_PHRASE_MARKERS):
        return True
    return "depth" in low and ("atoms/cc" in low or "atoms/cm" in low)


def _excel_preview_text(path: Path, max_rows: int = 8) -> str:
    """First few rows of sheet 0 flattened to a string (for content sniffing)."""
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = workbook.worksheets[0]
        cells: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            cells.extend(str(v) for v in row if v is not None)
    finally:
        workbook.close()
    return " ".join(cells)


def is_sims_file(path: Path) -> bool:
    """Content sniffer for ambiguous .csv/.tsv/.xlsx: True for a SIMS depth profile."""
    p = Path(path)
    try:
        if p.suffix.lower() in _EXCEL_EXTS:
            text = _excel_preview_text(p)
        else:
            text = p.read_text(encoding="latin-1", errors="replace")[:4096]
    except Exception:  # noqa: BLE001 — a sniffer must never raise; unreadable -> not SIMS
        return False
    return _sims_signals(text)


def _is_numeric(token: str) -> bool:
    try:
        value = float(token)
    except ValueError:
        return False
    return not math.isnan(value)


def _to_float(token: str) -> float:
    stripped = token.strip()
    if not stripped:
        return float("nan")
    try:
        return float(stripped)
    except ValueError:
        return float("nan")


def _read_raw_lines(text: str) -> list[str]:
    out: list[str] = []
    for raw in text.splitlines():
        stripped = raw.strip()
        if not stripped or stripped[0] in _COMMENT_CHARS:
            continue
        out.append(stripped)
    return out


def _detect_delimiter(raw_lines: Sequence[str]) -> str:
    test = raw_lines[:10]
    best, best_score = ",", 0.0
    for ch in _DELIM_CANDIDATES:
        counts = [line.count(ch) for line in test]
        if counts and all(c > 0 for c in counts):
            mean = sum(counts) / len(counts)
            std = (sum((c - mean) ** 2 for c in counts) / len(counts)) ** 0.5
            if std < mean * 0.5 and mean > best_score:
                best, best_score = ch, mean
    return best


def _numeric_score(row: Sequence[str]) -> float:
    if not row:
        return 0.0
    return sum(1 for t in row if _is_numeric(t.strip())) / len(row)


def _detect_layout(tokens: Sequence[Sequence[str]]) -> tuple[int, int]:
    """0-based (header_row, data_start); header walks back past blank rows."""
    scores = [_numeric_score(r) for r in tokens]
    first_data = next((i for i, s in enumerate(scores) if s > 0.5), 0)
    header_row = -1
    if first_data > 0:
        cand = first_data - 1
        while cand >= 0:
            if any(t.strip() for t in tokens[cand]) and scores[cand] < 0.5:
                break
            cand -= 1
        header_row = cand
    return header_row, first_data


def _detect_paired(matrix: np.ndarray) -> bool:
    n_cols = int(matrix.shape[1])
    if n_cols < 4 or n_cols % 2 != 0:
        return False
    n_odd = n_cols // 2
    n_monotonic = 0
    for k in range(n_odd):
        col = matrix[:, 2 * k]  # 0-based even = MATLAB odd (depth) columns
        col = col[~np.isnan(col)]
        if col.size >= 2 and bool(np.all(np.diff(col) > 0)):
            n_monotonic += 1
    return bool((n_monotonic / n_odd) >= 0.8)


_BARE_UNIT_RE = re.compile(r"^\(([^)]+)\)$")
_PAREN_RE = re.compile(r"(.+?)\s*\(([^)]+)\)\s*$")
_BRACK_RE = re.compile(r"(.+?)\s*\[([^\]]+)\]\s*$")
_CONC_RE = re.compile(r"^\s*(?:conc(?:entration)?)\s+", re.IGNORECASE)
_MASS_PREFIX_RE = re.compile(r"^\d+([A-Z][a-z]?)")
_MASS_TRAIL_RE = re.compile(r"^([A-Z][a-z]?)\d+[+\-]?$")


def _clean_element_names(headers: Sequence[str]) -> tuple[list[str], list[str]]:
    names: list[str] = []
    units: list[str] = []
    for header in headers:
        h = header.strip()
        bare = _BARE_UNIT_RE.match(h)
        if bare:
            names.append("")
            units.append(bare.group(1).strip())
            continue
        unit = ""
        paren = _PAREN_RE.match(h)
        brack = _BRACK_RE.match(h)
        if paren:
            h, unit = paren.group(1).strip(), paren.group(2).strip()
        elif brack:
            h, unit = brack.group(1).strip(), brack.group(2).strip()
        h = _CONC_RE.sub("", h)
        mass_prefix = _MASS_PREFIX_RE.match(h)
        if mass_prefix:
            h = mass_prefix.group(1)
        mass_trail = _MASS_TRAIL_RE.match(h)
        if mass_trail:
            h = mass_trail.group(1)
        h = re.sub(r"[+\-]+$", "", h)
        names.append(h)
        units.append(unit)
    return names, units


def _clean_vendor_element(raw: str) -> str:
    name = re.sub(r"-+>$", "", raw).strip()
    if name and name.isalpha():
        name = name[0].upper() + name[1:].lower()
    return name


def _matlab_colon(a: float, d: float, b: float) -> np.ndarray:
    n = int(math.floor((b - a) / d + 1e-10))
    grid = a + np.arange(n + 1) * d
    if grid[-1] < b:
        grid = np.append(grid, b)
    return grid


def _build_union_grid(
    depths: Sequence[np.ndarray], concs: Sequence[np.ndarray]
) -> tuple[np.ndarray, np.ndarray]:
    n_e = len(depths)
    all_same = True
    if n_e > 1:
        ref = depths[0]
        scale = float(np.max(np.abs(ref))) if ref.size else 0.0
        for e in range(1, n_e):
            if depths[e].size != ref.size or np.any(
                np.abs(depths[e] - ref) > np.spacing(scale) * 10
            ):
                all_same = False
                break
    if all_same and depths[0].size:
        union = depths[0].copy()
        return union, np.column_stack([c for c in concs])

    all_min, all_max, min_step = math.inf, -math.inf, math.inf
    for d in depths:
        if d.size == 0:
            continue
        all_min = min(all_min, float(d.min()))
        all_max = max(all_max, float(d.max()))
        steps = np.diff(d)
        pos = steps[steps > 0]
        if pos.size:
            min_step = min(min_step, float(pos.min()))

    if math.isinf(min_step) or all_min >= all_max:
        return depths[0].copy(), np.column_stack([c for c in concs])

    union = _matlab_colon(all_min, min_step, all_max)
    interp = np.full((union.size, n_e), np.nan)
    for e in range(n_e):
        d, c = depths[e], concs[e]
        if d.size < 2:
            if d.size:
                interp[int(np.argmin(np.abs(union - d[0]))), e] = c[0]
            continue
        interp[:, e] = np.interp(union, d, c, left=np.nan, right=np.nan)
    return union, interp


def _detect_depth_unit(col_headers: Sequence[str], header_meta: Sequence[str]) -> str:
    text = " ".join(list(col_headers) + list(header_meta)).lower()
    if "um" in text or "µ" in text or "micron" in text or "micrometer" in text:
        return "um"
    if "nm" in text or "nanometer" in text:
        return "nm"
    if "angstrom" in text or "Å" in " ".join(col_headers):
        return "A"
    return "nm"


def _read_text_tokens(path: Path) -> list[list[str]]:
    raw_lines = _read_raw_lines(path.read_text(encoding="latin-1"))
    if not raw_lines:
        raise ValueError(f"file empty or only comments: {path.name}")
    delim = _detect_delimiter(raw_lines)
    return [line.split(delim) for line in raw_lines]


def _read_excel_tokens(path: Path, sheet: int | str) -> list[list[str]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = workbook[sheet] if isinstance(sheet, str) else workbook.worksheets[sheet]
        grid = [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        workbook.close()
    tokens: list[list[str]] = []
    for row in grid:
        cells: list[str] = []
        for v in row:
            if isinstance(v, str):
                cells.append(v)
            elif isinstance(v, bool) or v is None:
                cells.append("")
            elif isinstance(v, (int, float)):
                cells.append(f"{v:.15g}")
            else:
                cells.append(str(v))
        tokens.append(cells)
    return tokens


def import_sims(
    filepath: str | Path,
    *,
    depth_unit: str = "auto",
    sheet: int | str = 0,
) -> DataStruct:
    """Import a SIMS depth profile (paired or shared-depth layout)."""
    path = Path(filepath)
    is_excel = path.suffix.lower() in {".xlsx", ".xls", ".xlsm", ".xlsb", ".ods"}
    tokens = _read_excel_tokens(path, sheet) if is_excel else _read_text_tokens(path)

    header_row, data_start = _detect_layout(tokens)
    data_rows = tokens[data_start:]
    n_data_cols = max(len(r) for r in data_rows)
    if header_row >= 0:
        col_headers = [c.strip() for c in tokens[header_row]]
    else:
        col_headers = []
    if len(col_headers) < n_data_cols:
        col_headers += [f"Col{k + 1}" for k in range(len(col_headers), n_data_cols)]
    elif len(col_headers) > n_data_cols:
        col_headers = col_headers[:n_data_cols]
    col_headers = [h if h.strip() else f"Col{k + 1}" for k, h in enumerate(col_headers)]

    header_meta: list[str] = []
    if header_row > 0:
        header_meta = [" ".join(t.strip() for t in tokens[mi]) for mi in range(header_row)]

    n_cols = len(col_headers)
    matrix = np.full((len(data_rows), n_cols), np.nan)
    for r, row in enumerate(data_rows):
        for c in range(min(len(row), n_cols)):
            matrix[r, c] = _to_float(row[c])

    empty_mask = np.all(np.isnan(matrix), axis=0)
    matrix = matrix[:, ~empty_mask]
    col_headers = [h for h, drop in zip(col_headers, empty_mask, strict=True) if not drop]
    n_cols = matrix.shape[1]
    if n_cols < 2:
        raise ValueError(f"need >=2 non-empty columns in {path.name}")

    is_paired = _detect_paired(matrix)
    depths: list[np.ndarray] = []
    concs: list[np.ndarray] = []
    elem_headers: list[str] = []
    if is_paired:
        for p in range(n_cols // 2):
            d_col, c_col = matrix[:, 2 * p], matrix[:, 2 * p + 1]
            valid = ~np.isnan(d_col) & ~np.isnan(c_col)
            depths.append(d_col[valid])
            concs.append(c_col[valid])
            elem_headers.append(col_headers[2 * p + 1])
    else:
        shared = matrix[:, 0]
        for e in range(1, n_cols):
            c_col = matrix[:, e]
            valid = ~np.isnan(shared) & ~np.isnan(c_col)
            depths.append(shared[valid])
            concs.append(c_col[valid])
            elem_headers.append(col_headers[e])

    elem_names, conc_units = _clean_element_names(elem_headers)
    if is_paired and any(not n for n in elem_names) and header_row > 0:
        _recover_paired_names(elem_names, tokens, header_row, empty_mask)

    union_depth, interp = _build_union_grid(depths, concs)
    if depth_unit != "auto":
        resolved_unit = depth_unit
    else:
        resolved_unit = _detect_depth_unit(col_headers, header_meta)

    metadata: dict[str, Any] = {
        "source": str(path),
        "parser_name": "import_sims",
        "x_column_name": "Depth",
        "x_column_unit": resolved_unit,
        "is_paired_layout": is_paired,
    }
    return DataStruct.create(
        union_depth, interp, labels=elem_names, units=conc_units, metadata=metadata
    )


def _recover_paired_names(
    elem_names: list[str],
    tokens: Sequence[Sequence[str]],
    header_row: int,
    empty_mask: np.ndarray,
) -> None:
    n_e = len(elem_names)
    width = empty_mask.size
    for mi in range(header_row - 1, -1, -1):
        row = list(tokens[mi])
        row = (row + [""] * width)[:width]
        parts = [cell for cell, drop in zip(row, empty_mask, strict=True) if not drop]
        parts = (parts + [""] * (2 * n_e))[: 2 * n_e]
        odd = [p.strip() for p in parts[0::2]]
        even = [p.strip() for p in parts[1::2]]
        n_even_blank = sum(1 for p in even if not p)
        n_odd_text = sum(1 for p in odd if p and not _is_numeric(p))
        if n_even_blank >= n_e // 2 and n_odd_text >= n_e // 2:
            for p in range(n_e):
                if not elem_names[p] and odd[p]:
                    elem_names[p] = _clean_vendor_element(odd[p])
            return
